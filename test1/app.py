from __future__ import annotations

import argparse
import csv
import io
import math
import os
import re
import sqlite3
from dataclasses import dataclass
from datetime import datetime
from typing import Any, Dict, Optional, Tuple, List

from flask import Flask, flash, redirect, render_template, request, session, url_for, Response


BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.path.join(BASE_DIR, "students.db")

ADMIN_USERNAME = os.environ.get("SMS_ADMIN_USERNAME", "admin")
ADMIN_PASSWORD = os.environ.get("SMS_ADMIN_PASSWORD", "admin123")


def get_conn() -> sqlite3.Connection:
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    return conn


def init_db() -> None:
    with get_conn() as conn:
        conn.executescript(
            """
            CREATE TABLE IF NOT EXISTS students (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                student_no TEXT NOT NULL UNIQUE,
                name TEXT NOT NULL,
                gender TEXT NOT NULL DEFAULT '未知',
                age INTEGER,
                class_name TEXT,
                phone TEXT,
                note TEXT,
                created_at TEXT NOT NULL
            );

            CREATE INDEX IF NOT EXISTS idx_students_name ON students(name);
            CREATE INDEX IF NOT EXISTS idx_students_student_no ON students(student_no);
            CREATE INDEX IF NOT EXISTS idx_students_class_name ON students(class_name);
            """
        )


@dataclass
class StudentInput:
    student_no: str
    name: str
    gender: str
    age: Optional[int]
    class_name: str
    phone: str
    note: str


def _clean_str(v: Any) -> str:
    if v is None:
        return ""
    return str(v).strip()


def validate_student_form(form: Dict[str, Any]) -> Tuple[Optional[StudentInput], Dict[str, str]]:
    errors: Dict[str, str] = {}
    student_no = _clean_str(form.get("student_no"))
    name = _clean_str(form.get("name"))
    gender = _clean_str(form.get("gender")) or "未知"
    age_raw = _clean_str(form.get("age"))
    class_name = _clean_str(form.get("class_name"))
    phone = _clean_str(form.get("phone"))
    note = _clean_str(form.get("note"))

    if not student_no:
        errors["student_no"] = "学号不能为空"
    elif not re.fullmatch(r"[A-Za-z0-9_-]{4,32}", student_no):
        errors["student_no"] = "学号格式不正确（建议 4-32 位字母/数字/下划线/短横线）"

    if not name:
        errors["name"] = "姓名不能为空"
    elif len(name) > 40:
        errors["name"] = "姓名过长（最多 40 字）"

    if gender not in {"男", "女", "未知"}:
        errors["gender"] = "性别只能选：男/女/未知"

    age: Optional[int] = None
    if age_raw:
        if not age_raw.isdigit():
            errors["age"] = "年龄必须是数字"
        else:
            age = int(age_raw)
            if age < 1 or age > 120:
                errors["age"] = "年龄范围应为 1-120"

    if len(class_name) > 40:
        errors["class_name"] = "班级过长（最多 40 字）"

    if phone:
        if len(phone) > 30:
            errors["phone"] = "电话过长（最多 30 字）"
        elif not re.fullmatch(r"[0-9+() -]{5,30}", phone):
            errors["phone"] = "电话格式不正确"

    if len(note) > 200:
        errors["note"] = "备注过长（最多 200 字）"

    if errors:
        return None, errors

    return (
        StudentInput(
            student_no=student_no,
            name=name,
            gender=gender,
            age=age,
            class_name=class_name,
            phone=phone,
            note=note,
        ),
        {},
    )


def create_app() -> Flask:
    app = Flask(__name__)
    app.secret_key = os.environ.get("FLASK_SECRET_KEY", "dev-secret-key-change-me")

    def is_logged_in() -> bool:
        return session.get("admin_logged_in") is True

    @app.before_request
    def require_login():
        # 允许访问登录页、静态资源、favicon
        if request.endpoint in {"login", "do_login", "static"}:
            return None
        if request.path.startswith("/static/") or request.path == "/favicon.ico":
            return None
        if is_logged_in():
            return None
        return redirect(url_for("login", next=request.full_path))

    @app.get("/login")
    def login():
        if is_logged_in():
            return redirect(url_for("index"))
        next_url = _clean_str(request.args.get("next")) or ""
        return render_template("login.html", next=next_url)

    @app.get("/login.html")
    def login_html():
        # 兼容用户手动输入 /login.html
        return redirect(url_for("login", next=request.args.get("next", "")))

    @app.post("/login")
    def do_login():
        username = _clean_str(request.form.get("username"))
        password = _clean_str(request.form.get("password"))
        next_url = _clean_str(request.form.get("next")) or ""

        if username == ADMIN_USERNAME and password == ADMIN_PASSWORD:
            session["admin_logged_in"] = True
            session["admin_username"] = username
            flash("登录成功。", "success")
            # 简单安全：只允许站内跳转
            if next_url.startswith("/") and not next_url.startswith("//"):
                return redirect(next_url)
            return redirect(url_for("index"))

        flash("账号或密码错误。", "error")
        return render_template("login.html", next=next_url), 401

    @app.post("/logout")
    def logout():
        session.clear()
        flash("已退出登录。", "success")
        return redirect(url_for("login"))

    def _normalize_header(h: str) -> str:
        h = _clean_str(h).lower()
        h = h.replace(" ", "").replace("\ufeff", "")
        mapping = {
            "学号": "student_no",
            "studentno": "student_no",
            "student_no": "student_no",
            "学号/工号": "student_no",
            "姓名": "name",
            "name": "name",
            "性别": "gender",
            "gender": "gender",
            "年龄": "age",
            "age": "age",
            "班级": "class_name",
            "class": "class_name",
            "classname": "class_name",
            "class_name": "class_name",
            "电话": "phone",
            "手机号": "phone",
            "phone": "phone",
            "备注": "note",
            "note": "note",
        }
        return mapping.get(h, h)

    def _decode_upload(data: bytes) -> str:
        # 尽量兼容常见编码：UTF-8(BOM)/UTF-8/GBK
        for enc in ("utf-8-sig", "utf-8", "gbk"):
            try:
                return data.decode(enc)
            except UnicodeDecodeError:
                continue
        return data.decode("utf-8", errors="replace")

    def _read_csv_rows(file_bytes: bytes) -> Tuple[List[str], List[Dict[str, Any]], str]:
        text = _decode_upload(file_bytes)
        if not text.strip():
            raise ValueError("文件内容为空")

        sample = text[:4096]
        delimiter = "," if sample.count(",") >= sample.count("\t") else "\t"
        reader = csv.DictReader(io.StringIO(text), delimiter=delimiter)
        if reader.fieldnames is None:
            raise ValueError("无法识别表头，请确认第一行是列名")
        fieldnames = list(reader.fieldnames)
        rows = [dict(r) for r in reader]
        human_delim = "逗号(,)" if delimiter == "," else "制表符(\\t)"
        return fieldnames, rows, human_delim

    def _read_xlsx_rows(file_bytes: bytes) -> Tuple[List[str], List[Dict[str, Any]], str]:
        try:
            from openpyxl import load_workbook  # type: ignore
        except Exception as e:  # pragma: no cover
            raise RuntimeError("缺少 Excel 解析依赖 openpyxl，请先安装依赖后重试") from e

        wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        ws = wb.worksheets[0]
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration:
            raise ValueError("Excel 内容为空")

        fieldnames: List[str] = [_clean_str(x) for x in header_row]
        if not any(fieldnames):
            raise ValueError("无法识别表头，请确认第一行是列名")

        rows: List[Dict[str, Any]] = []
        for values in rows_iter:
            if values is None:
                continue
            # 空行跳过
            if all(_clean_str(v) == "" for v in values):
                continue
            row: Dict[str, Any] = {}
            for i, key in enumerate(fieldnames):
                if not key:
                    continue
                cell = values[i] if i < len(values) else ""
                row[key] = "" if cell is None else str(cell)
            rows.append(row)

        return fieldnames, rows, "Excel(.xlsx)"

    @app.get("/students/import")
    def import_students_page():
        return render_template("import.html", result=None)

    @app.post("/students/import")
    def import_students():
        file = request.files.get("file")
        mode = _clean_str(request.form.get("mode")) or "upsert"
        if file is None or not file.filename:
            flash("请选择要导入的 CSV 文件。", "error")
            return render_template("import.html", result=None), 400

        raw = file.read()
        filename = _clean_str(file.filename).lower()

        try:
            if filename.endswith(".xlsx"):
                fieldnames, raw_rows, source_type = _read_xlsx_rows(raw)
            else:
                fieldnames, raw_rows, source_type = _read_csv_rows(raw)
        except Exception as e:
            flash(f"解析文件失败：{e}", "error")
            return render_template("import.html", result=None), 400

        normalized_fieldnames = [_normalize_header(h) for h in fieldnames]

        rows_to_insert = []
        now = datetime.now().isoformat(timespec="seconds")
        errors = []
        total = len(raw_rows)

        # 行号：CSV 为真实行号（含表头），XLSX 这里按“数据行从 2 开始”展示
        for offset, row in enumerate(raw_rows, start=0):
            idx = 2 + offset
            normalized: Dict[str, Any] = {}
            for k, v in (row or {}).items():
                if k is None:
                    continue
                nk = _normalize_header(k)
                normalized[nk] = v

            data, e = validate_student_form(normalized)
            if e or data is None:
                errors.append({"line": idx, "student_no": _clean_str(normalized.get("student_no")), "errors": e})
                continue

            rows_to_insert.append(
                (
                    data.student_no,
                    data.name,
                    data.gender,
                    data.age,
                    data.class_name,
                    data.phone,
                    data.note,
                    now,
                )
            )

        inserted = 0
        updated = 0

        if rows_to_insert:
            with get_conn() as conn:
                if mode not in {"append", "upsert"}:
                    mode = "upsert"

                if mode == "append":
                    try:
                        conn.executemany(
                            """
                            INSERT INTO students (student_no, name, gender, age, class_name, phone, note, created_at)
                            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                            """,
                            rows_to_insert,
                        )
                        inserted = len(rows_to_insert)
                    except sqlite3.IntegrityError:
                        # 如果 append 模式遇到重复学号，逐行插入以定位
                        for r in rows_to_insert:
                            try:
                                conn.execute(
                                    """
                                    INSERT INTO students (student_no, name, gender, age, class_name, phone, note, created_at)
                                    VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                                    """,
                                    r,
                                )
                                inserted += 1
                            except sqlite3.IntegrityError:
                                errors.append(
                                    {
                                        "line": None,
                                        "student_no": r[0],
                                        "errors": {"student_no": "学号已存在（append 模式不允许覆盖）"},
                                    }
                                )
                else:
                    # upsert：学号存在则更新（不改 created_at）
                    conn.executemany(
                        """
                        INSERT INTO students (student_no, name, gender, age, class_name, phone, note, created_at)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                        ON CONFLICT(student_no) DO UPDATE SET
                          name = excluded.name,
                          gender = excluded.gender,
                          age = excluded.age,
                          class_name = excluded.class_name,
                          phone = excluded.phone,
                          note = excluded.note
                        """,
                        rows_to_insert,
                    )
                    # SQLite 无法直接区分 insert/update 行数，这里用“可用行数”展示
                    inserted = len(rows_to_insert)

        result = {
            "filename": file.filename,
            "delimiter": source_type,
            "fieldnames": normalized_fieldnames,
            "total_rows": total,
            "accepted_rows": len(rows_to_insert),
            "inserted": inserted,
            "updated": updated,
            "error_count": len(errors),
            "errors": errors[:50],  # 页面只展示前 50 条
            "mode": mode,
        }

        if errors:
            flash(f"导入完成：成功 {len(rows_to_insert)} 行，失败 {len(errors)} 行（仅展示前 50 条错误）。", "error")
        else:
            flash(f"导入成功：共写入 {len(rows_to_insert)} 行。", "success")

        return render_template("import.html", result=result)

    @app.get("/students/import/sample.csv")
    def import_sample_csv():
        content = "学号,姓名,性别,年龄,班级,电话,备注\n20260001,张三,男,20,计科1班,13800000000,示例数据\n"
        return Response(
            content,
            mimetype="text/csv; charset=utf-8",
            headers={"Content-Disposition": "attachment; filename=students_sample.csv"},
        )

    @app.get("/")
    def index():
        q = _clean_str(request.args.get("q"))
        page = _clean_str(request.args.get("page")) or "1"
        per_page = 10

        try:
            page_i = max(1, int(page))
        except ValueError:
            page_i = 1

        where = ""
        params: Tuple[Any, ...] = ()
        if q:
            where = "WHERE student_no LIKE ? OR name LIKE ? OR class_name LIKE ?"
            like = f"%{q}%"
            params = (like, like, like)

        with get_conn() as conn:
            total = conn.execute(f"SELECT COUNT(1) AS c FROM students {where}", params).fetchone()["c"]
            pages = max(1, math.ceil(total / per_page)) if total else 1
            page_i = min(page_i, pages)
            offset = (page_i - 1) * per_page

            rows = conn.execute(
                f"""
                SELECT id, student_no, name, gender, age, class_name, phone, note, created_at
                FROM students
                {where}
                ORDER BY datetime(created_at) DESC, id DESC
                LIMIT ? OFFSET ?
                """,
                params + (per_page, offset),
            ).fetchall()

        return render_template(
            "index.html",
            rows=rows,
            q=q,
            page=page_i,
            pages=pages,
            total=total,
            per_page=per_page,
        )

    @app.get("/index")
    @app.get("/index.html")
    def index_alias():
        # 兼容 /index 与 /index.html
        return redirect(url_for("index"))

    @app.get("/students/new")
    def new_student():
        return render_template("form.html", mode="create", student=None, errors={})

    @app.post("/students")
    def create_student():
        data, errors = validate_student_form(request.form)
        if errors or data is None:
            flash("请先修正表单错误后再提交。", "error")
            return render_template("form.html", mode="create", student=request.form, errors=errors), 400

        try:
            with get_conn() as conn:
                conn.execute(
                    """
                    INSERT INTO students (student_no, name, gender, age, class_name, phone, note, created_at)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        data.student_no,
                        data.name,
                        data.gender,
                        data.age,
                        data.class_name,
                        data.phone,
                        data.note,
                        datetime.now().isoformat(timespec="seconds"),
                    ),
                )
            flash("新增学生成功。", "success")
            return redirect(url_for("index"))
        except sqlite3.IntegrityError:
            flash("学号已存在，请更换后再试。", "error")
            return render_template("form.html", mode="create", student=request.form, errors={"student_no": "学号已存在"}), 409

    def _get_student_or_404(student_id: int):
        with get_conn() as conn:
            row = conn.execute(
                """
                SELECT id, student_no, name, gender, age, class_name, phone, note, created_at
                FROM students
                WHERE id = ?
                """,
                (student_id,),
            ).fetchone()
        if row is None:
            return None
        return row

    @app.get("/students/<int:student_id>/edit")
    def edit_student(student_id: int):
        row = _get_student_or_404(student_id)
        if row is None:
            flash("未找到该学生记录。", "error")
            return redirect(url_for("index")), 404
        return render_template("form.html", mode="edit", student=row, errors={})

    @app.post("/students/<int:student_id>")
    def update_student(student_id: int):
        existing = _get_student_or_404(student_id)
        if existing is None:
            flash("未找到该学生记录。", "error")
            return redirect(url_for("index")), 404

        data, errors = validate_student_form(request.form)
        if errors or data is None:
            flash("请先修正表单错误后再提交。", "error")
            return render_template("form.html", mode="edit", student={**dict(existing), **request.form}, errors=errors), 400

        try:
            with get_conn() as conn:
                conn.execute(
                    """
                    UPDATE students
                    SET student_no = ?, name = ?, gender = ?, age = ?, class_name = ?, phone = ?, note = ?
                    WHERE id = ?
                    """,
                    (
                        data.student_no,
                        data.name,
                        data.gender,
                        data.age,
                        data.class_name,
                        data.phone,
                        data.note,
                        student_id,
                    ),
                )
            flash("更新成功。", "success")
            return redirect(url_for("index"))
        except sqlite3.IntegrityError:
            flash("学号已存在，请更换后再试。", "error")
            return (
                render_template(
                    "form.html",
                    mode="edit",
                    student={**dict(existing), **request.form},
                    errors={"student_no": "学号已存在"},
                ),
                409,
            )

    @app.post("/students/<int:student_id>/delete")
    def delete_student(student_id: int):
        with get_conn() as conn:
            cur = conn.execute("DELETE FROM students WHERE id = ?", (student_id,))
            deleted = cur.rowcount
        if deleted:
            flash("删除成功。", "success")
        else:
            flash("未找到该学生记录。", "error")
        return redirect(url_for("index"))

    return app


def main() -> None:
    parser = argparse.ArgumentParser(description="学生管理系统（Flask + SQLite）")
    parser.add_argument("--init-db", action="store_true", help="初始化数据库表结构")
    parser.add_argument("--host", default="127.0.0.1")
    parser.add_argument("--port", default=5000, type=int)
    parser.add_argument("--debug", action="store_true")
    args = parser.parse_args()

    if args.init_db:
        init_db()
        print(f"数据库已初始化：{DB_PATH}")
        return

    init_db()
    app = create_app()
    app.run(host=args.host, port=args.port, debug=args.debug)


if __name__ == "__main__":
    main()

